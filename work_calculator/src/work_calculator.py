from datetime import datetime
import string

import pandas as pd
from openpyxl import load_workbook
from tabulate import tabulate

from work_calculator.src.util import match_string, transform_to_date, get_name, get_last_name
from work_calculator.src.work_calculator import *

class WorkCalculator:
    date = datetime.today().strftime('%Y-%m-%d-%H-%M')
    columns = ['ay', 'çalışan', 'toplam_çalışma_süresi', 'toplam_yolda_geçen_süre', 'toplam_süre']
    output_df = pd.DataFrame(columns=columns)
    font_size = '14pt;'
    EXTENDED_UPPERCASE_ALPHABET = list(string.ascii_uppercase) + ['A' + i for i in string.ascii_uppercase]
    ignore_tur_strings = [
        "urlaub",
        "dienstbesprechung",
        "arztbesuch",
        "büro"
    ]
    drop_row_strings = [
        "freie Wochenende arbeiten"
    ]

    def __init__(self,
                 input_path,
                 output_path):
        self.input_path = input_path
        self.output_path = self.date + "-report.xlsx" if not output_path else output_path
        self.workbook = load_workbook(input_path)

    def add_tur_ignore_check(self, df):
        df['ignore_tur'] = df['hasta'].apply(lambda x: match_string(x, self.ignore_tur_strings))
        return df

    def add_row_drop_check(self, df):
        df['drop_row'] = df['hasta'].apply(lambda x: match_string(x, self.drop_row_strings))
        return df

    @staticmethod
    def add_tur_sayisi(df):
        grped = df.groupby(['tarih', 'çalışan'])
        new_df = pd.DataFrame(columns=['tarih',
                                       'dakika',
                                       'hasta',
                                       'çalışan',
                                       'hasta_ad',
                                       'hasta_soyad',
                                       'günlük_tur_sayısı',
                                       'günlük_çalışma',
                                       'yolda_geçen_süre'])

        for group_name, grp in grped:
            ignore_count = len(grp[grp["ignore_tur"] == True])
            for index, row in grp.iterrows():
                row['günlük_tur_sayısı'] = len(grp)-ignore_count
                row['günlük_çalışma'] = grp['dakika'].sum()
                if row['günlük_tur_sayısı'] >= 2:
                    row['yolda_geçen_süre'] = (row['günlük_tur_sayısı'] - 1) * 15
                else:
                    row['yolda_geçen_süre'] = 0
                new_df.loc[len(new_df)] = row
        return new_df

    @staticmethod
    def process_df(df, sheetname):
        df = df[['tarih', 'çalışan', 'günlük_tur_sayısı', 'günlük_çalışma', 'yolda_geçen_süre']].drop_duplicates()
        df['toplam_çalışma_süresi'] = df['günlük_çalışma'].groupby(df['çalışan']).transform('sum')
        df['toplam_tur_sayısı'] = df['günlük_tur_sayısı'].groupby(df['çalışan']).transform('sum')
        df['toplam_yolda_geçen_süre'] = df['yolda_geçen_süre'].groupby(df['çalışan']).transform('sum')
        df['toplam_süre'] = df['toplam_çalışma_süresi'] + df['toplam_yolda_geçen_süre']
        df['toplam_süre'] = df['toplam_süre'].apply(transform_to_date)
        df = df.sort_values('toplam_çalışma_süresi', ascending=False)
        df['ay'] = sheetname
        df = df[['ay', 'çalışan', 'toplam_çalışma_süresi', 'toplam_yolda_geçen_süre', 'toplam_süre']]
        df = df.drop_duplicates()
        empty_row = pd.DataFrame(
            {"ay": None, "çalışan": None, 'toplam_çalışma_süresi': None, 'toplam_yolda_geçen_süre': None,
             'toplam_süre': None}, index=[0])
        df = pd.concat([empty_row, df.loc[:]]).reset_index(drop=True)
        return df

    @staticmethod
    def clean_df(df2):
        df = df2.copy()
        df = df[[2, 4, 6, 7]]
        df.columns = ["tarih", "dakika", "hasta", 'çalışan']
        df["tarih"] = df['tarih'].apply(lambda x: x[1:-1])
        df["dakika"] = df['dakika'].apply(lambda x: int(x[1:-1]))
        df["çalışan"] = df['çalışan'].apply(lambda x: x[1:-1])
        df["hasta"] = df['hasta'].apply(lambda x: x[1:-1])
        df["hasta_ad"] = df['hasta'].apply(get_name)
        df["hasta_soyad"] = df['hasta'].apply(get_last_name)
        return df

    def save_report(self):
        self.output_df = self.output_df[
            ['ay', 'çalışan', 'toplam_çalışma_süresi', 'toplam_yolda_geçen_süre', 'toplam_süre']]
        # self.output_df.to_excel(self.output_path, engine="openpyxl", index=False)

    def prepare_df(self, sheet_name):
        df = pd.read_excel(self.input_path, engine="openpyxl", sheet_name=sheet_name, header=None)
        df = self.clean_df(df)
        df = self.add_tur_ignore_check(df)
        df = self.add_row_drop_check(df)
        print(tabulate(df, headers="keys"))
        df = self.add_tur_sayisi(df)
        df = df[['tarih', 'çalışan', 'günlük_tur_sayısı', 'günlük_çalışma', 'yolda_geçen_süre']].drop_duplicates()
        return df

    def create_report(self):
        for sheet in self.workbook.worksheets:
            df = self.prepare_df(sheet.title)
            df = self.process_df(df, sheet.title)
            self.output_df = pd.concat([self.output_df, df])
        self.save_report()
        self.add_detailed_daily()
        self.format_report()
        self.adjust_width()

    def adjust_width(self, path=None):
        from openpyxl import load_workbook
        if path:
            path = path
        else:
            path = self.output_path

        wb = load_workbook(path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")

            excel_keymap = {}
            for i, column in enumerate(df.columns):
                excel_keymap[column] = self.EXTENDED_UPPERCASE_ALPHABET[i]
            # Auto-adjust columns' width
            for column in df.columns:
                column_width = max(df[column].astype(str).map(len).max(), len(column))
                sheet.column_dimensions[excel_keymap[column]].width = column_width - 1
        wb.save(path)

    def adjust_font(self, x):
        return f"font-size: {self.font_size}; font-weight: bold"

    def format_report(self):
        wb = load_workbook(self.output_path)

        for sheet_name in wb.sheetnames:
            df = pd.read_excel(self.output_path, engine='openpyxl', sheet_name=sheet_name)
            writer = pd.ExcelWriter(self.output_path, engine='openpyxl', mode='a')
            df = df.style.applymap(self.adjust_font)
            writer.book = wb
            wb.remove(wb[sheet_name])
            df.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
            writer.close()

    def add_detailed_daily(self):
        writer = pd.ExcelWriter(self.output_path, engine='openpyxl', mode='a')
        for sheet in self.workbook.worksheets:
            df = self.prepare_df(sheet.title)
            df.to_excel(excel_writer=writer, sheet_name=f"daily-detailed-{sheet.title}", index=False)
        writer.close()